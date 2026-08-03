# -*- coding: utf-8 -*-
"""教材加值產生引擎：吃一個 lesson dict，產出 摘要MD / 簡報HTML / 圖表HTML / Bloom MD / 測驗XLSX"""
import os, importlib.util
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "/mnt/user-data/outputs"

_gen = os.path.dirname(os.path.abspath(__file__))
_s = importlib.util.spec_from_file_location('mcq', os.path.join(_gen, 'mcq.py'))
mcq = importlib.util.module_from_spec(_s); _s.loader.exec_module(mcq)

# ---------- 1. 摘要 MD ----------
def gen_summary(L):
    themes = "\n\n".join(
        f"### 主題{i+1}：{t['name']}\n" + "\n".join(f"- {p}" for p in t['points'])
        for i, t in enumerate(L['themes'])
    )
    misc = "\n".join(f"| {m[0]} | {m[1]} |" for m in L['misconceptions'])
    life = "\n".join(f"- **{x[0]}**：{x[1]}" for x in L['life'])
    md = f"""# {L['grade']} {L['code']}《{L['title']}》— 教材摘要與重點主題

## 一、本節摘要

{L['summary']}

## 二、{len(L['themes'])}大重點主題

{themes}

## 三、常見迷思與教學提醒

| 迷思 | 正確觀念 |
|---|---|
{misc}

## 四、生活與職場連結（素養導向）
{life}
"""
    p = os.path.join(OUT, f"{L['code']}_教材摘要與重點主題.md")
    open(p, "w", encoding="utf-8").write(md)
    return p

# ---------- 2. Bloom MD ----------
def gen_bloom(L):
    b = L['bloom']
    md = f"""# {L['grade']} {L['code']}《{L['title']}》— Bloom 認知層次差異化教材報告

> 依 Bloom 認知層次金字塔，將本節設計為三份差異化教材。教師可依班級狀況分組指派。

Bloom 六層次（由低到高）：**記憶 → 理解 → 應用 → 分析 → 評鑑 → 創造**

---

# 報告 A：基礎鞏固版（記憶・理解）
**適用對象**：需打穩基礎、概念仍陌生的學生

## 教學重點
{b['A']['focus']}

## 練習設計
{chr(10).join('- ' + x for x in b['A']['tasks'])}

**檢核標準**：{b['A']['check']}

---

# 報告 B：標準精熟版（應用・分析）
**適用對象**：已掌握基本概念、需熟練運算與比較的一般學生

## 教學重點
{b['B']['focus']}

## 練習設計
{chr(10).join('- ' + x for x in b['B']['tasks'])}

**檢核標準**：{b['B']['check']}

---

# 報告 C：挑戰延伸版（評鑑・創造）
**適用對象**：學有餘力、需要挑戰的學生

## 教學重點
{b['C']['focus']}

## 專題與挑戰
{chr(10).join('- ' + x for x in b['C']['tasks'])}

**檢核標準**：{b['C']['check']}

---

# 三版對照總表

| 面向 | A 基礎鞏固 | B 標準精熟 | C 挑戰延伸 |
|---|---|---|---|
| Bloom 層次 | 記憶・理解 | 應用・分析 | 評鑑・創造 |
| 核心目標 | 認得、說得出 | 會用、能比較 | 能判斷、能創造 |
| 支援程度 | 高 | 中 | 低 |

**教師使用建議**：可採「同課異步」，三組並行，最後全班分享 C 組創作題。
"""
    p = os.path.join(OUT, f"{L['code']}_Bloom差異化教材報告.md")
    open(p, "w", encoding="utf-8").write(md)
    return p

# ---------- 3. 測驗 XLSX ----------
def gen_quiz(L):
    title_font=Font(name='Arial',size=14,bold=True,color='FFFFFF')
    header_font=Font(name='Arial',size=11,bold=True,color='FFFFFF')
    cell_font=Font(name='Arial',size=11)
    ans_font=Font(name='Arial',size=11,color='C00000',bold=True)
    title_fill=PatternFill('solid',fgColor='1A2E1A')
    header_fill=PatternFill('solid',fgColor='4A7A4A')
    alt_fill=PatternFill('solid',fgColor='F0F5F0')
    thin=Side(style='thin',color='BBBBBB')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    wrap=Alignment(wrap_text=True,vertical='center')
    center=Alignment(horizontal='center',vertical='center')
    labels={'基礎卷':'★☆☆ 基礎（記憶・理解）','進階卷':'★★☆ 進階（應用・分析）','挑戰卷':'★★★ 挑戰（評鑑・創造）'}
    M = mcq.build(L['quiz'], L['code'])
    wb=Workbook(); first=True
    for sheet in L['quiz']:
        items=M.get(sheet, [])
        ws=wb.active if first else wb.create_sheet(sheet)
        if first: ws.title=sheet; first=False
        widths=[6,54,20,20,20,20,8,40]
        for col,w in zip('ABCDEFGH',widths): ws.column_dimensions[col].width=w
        ws.merge_cells('A1:H1')
        c=ws['A1']; c.value=f"{L['grade']} {L['code']} {L['title']}　測驗　{labels[sheet]}（四選一單選）"
        c.font=title_font; c.fill=title_fill; c.alignment=center
        ws.row_dimensions[1].height=30
        for i,h in enumerate(['題號','題目','(A)','(B)','(C)','(D)','答案','解析'],1):
            cell=ws.cell(row=2,column=i,value=h); cell.font=header_font
            cell.fill=header_fill; cell.alignment=center; cell.border=border
        ws.row_dimensions[2].height=22
        for idx,it in enumerate(items,1):
            r=idx+2
            ws.cell(row=r,column=1,value=idx).alignment=center
            ws.cell(row=r,column=2,value=it['q']).alignment=wrap
            for k in range(4):
                oc=ws.cell(row=r,column=3+k,value=it['opts'][k]); oc.alignment=wrap
            ac=ws.cell(row=r,column=7,value=chr(65+it['ci'])); ac.alignment=center; ac.font=ans_font
            ws.cell(row=r,column=8,value=it['e']).alignment=wrap
            for col in range(1,9):
                cc=ws.cell(row=r,column=col); cc.border=border
                if col!=7: cc.font=cell_font
                if idx%2==0: cc.fill=alt_fill
            ws.row_dimensions[r].height=40
        ws.freeze_panes='A3'
    p=os.path.join(OUT,f"{L['code']}_三種難度測驗卷.xlsx")
    wb.save(p)
    return p

print("engine loaded")
