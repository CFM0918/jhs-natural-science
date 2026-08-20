# -*- coding: utf-8 -*-
"""為每個已完成的歷屆試題年份產生 XLSX(需 openpyxl，用 py 執行)。"""
import os, importlib.util
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

GEN = os.path.dirname(os.path.abspath(__file__)); ROOT = os.path.dirname(GEN)
SRC = os.path.join(GEN, '_exam_src')
OUTDIR = os.path.join(ROOT, '02_加值成品', '歷屆試題')

def load_exam(year):
    p = os.path.join(SRC, f'exam_{year}.py')
    if not os.path.exists(p): return None
    s = importlib.util.spec_from_file_location(f'exam_{year}', p)
    m = importlib.util.module_from_spec(s); s.loader.exec_module(m)
    return m.ITEMS

def gen_xlsx(year, items):
    title_font=Font(name='Arial',size=14,bold=True,color='FFFFFF')
    header_font=Font(name='Arial',size=11,bold=True,color='FFFFFF')
    cell_font=Font(name='Arial',size=11)
    ans_font=Font(name='Arial',size=11,color='C00000',bold=True)
    title_fill=PatternFill('solid',fgColor='1A2E1A')
    header_fill=PatternFill('solid',fgColor='4A7A4A')
    alt_fill=PatternFill('solid',fgColor='F0F5F0')
    thin=Side(style='thin',color='BBBBBB')
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    wrap=Alignment(wrap_text=True,vertical='center')
    center=Alignment(horizontal='center',vertical='center')
    wb=Workbook(); ws=wb.active; ws.title='自然科'
    widths=[6,50,20,20,20,20,8,42]
    for col,w in zip('ABCDEFGH',widths): ws.column_dimensions[col].width=w
    ws.merge_cells('A1:H1')
    c=ws['A1']; c.value=f'{year}年國中教育會考自然科試題（官方題目與正解・詳解為本站原創）'
    c.font=title_font; c.fill=title_fill; c.alignment=center
    ws.row_dimensions[1].height=30
    for i,h in enumerate(['題號','題目','(A)','(B)','(C)','(D)','答案','解析'],1):
        cell=ws.cell(row=2,column=i,value=h); cell.font=header_font
        cell.fill=header_fill; cell.alignment=center; cell.border=border
    ws.row_dimensions[2].height=22
    for idx,it in enumerate(items,1):
        r=idx+2
        ws.cell(row=r,column=1,value=it['n']).alignment=center
        ws.cell(row=r,column=2,value=it['stem']).alignment=wrap
        if it['opts']:
            for k,letter in enumerate('ABCD'):
                ws.cell(row=r,column=3+k,value=it['opts'][letter]).alignment=wrap
        else:
            for k in range(4):
                ws.cell(row=r,column=3+k,value='(圖形選項，請見官方題本PDF)').alignment=wrap
        ac=ws.cell(row=r,column=7,value=it['ans']); ac.alignment=center; ac.font=ans_font
        ws.cell(row=r,column=8,value=it['explain']).alignment=wrap
        for col in range(1,9):
            cc=ws.cell(row=r,column=col); cc.border=border
            if col!=7: cc.font=cell_font
            if idx%2==0: cc.fill=alt_fill
        ws.row_dimensions[r].height=44
    ws.freeze_panes='A3'
    out=os.path.join(OUTDIR, f'{year}年會考自然科.xlsx')
    wb.save(out)
    return out

def main():
    n=0
    for y in range(105,116):
        items=load_exam(y)
        if items is None: continue
        gen_xlsx(y, items); n+=1
    print(f'OK: 產出 {n} 份歷屆試題 XLSX')

if __name__=='__main__': main()
